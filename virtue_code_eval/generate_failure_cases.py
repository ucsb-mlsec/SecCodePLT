from pathlib import Path

import orjson
import pandas as pd
import logging
import argparse

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font

logger = logging.getLogger(__name__)


def adjust_excel(excel_path):
    # adjust the column width for the Excel file
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(name="roboto", size=12)
    # adjust the column width
    max_width = 60
    for col in sheet.iter_cols():
        max_length = 0
        percentage = 1  # Adjust the width of the column for float value
        column = None  # Get the column name
        for cell in col:
            if column is None and "column_letter" in dir(cell):
                column = cell.column_letter
            try:
                _ = float(cell.value)
                percentage = 1.3  # magic number for adjusting float value
            except (TypeError, ValueError):
                pass
            if "\n" in str(cell.value):
                max_length = max(
                    max_length, *[len(i) for i in str(cell.value).split("\n")]
                )
            else:
                max_length = max(max_length, len(str(cell.value)) / percentage)
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > max_width:
            adjusted_width = max_width
        sheet.column_dimensions[column].width = adjusted_width
    workbook.save(excel_path)


def generate_failure_cases(data_folder: Path, json_results_path: Path) -> None:
    """Generate a table from the results of the evaluation.
    :param data_folder: str
        The folder where the results will be saved.
    :param json_results_path: str
        The path to the json results file.
    """
    # read json file
    with open(json_results_path, "rb") as f:
        data = orjson.loads(f.read())
    # store data in a dictionary
    df_data = []
    # iterate over the data and store the metrics in the dictionary
    for model_name, model_data in data.items():
        for task_name, task_data in model_data.items():
            for each_data in task_data:
                subtasks = sorted(each_data["subtask"].values())
                subtasks_str = "_".join(subtasks)
                prompt = ""
                for message in each_data["messages"]:
                    role = message["role"]
                    role_prompt = message["content"]
                    prompt += role + ":\n" + role_prompt + "\n\n"
                prompt = prompt.strip()
                data_entry = {
                    "Model": model_name,
                    "Task": task_name,
                    "Subtask": subtasks_str,
                    "prompt": prompt,
                    "response": each_data["response"],
                    "reference": each_data["reference"],
                    "id": each_data["id_"],
                }
                # convert bool to float
                each_data["metrics"] = { # convert bool to float
                    key: float(value) if isinstance(value, bool) else value
                    for key, value in each_data["metrics"].items()
                }
                data_entry.update(each_data["metrics"])
                df_data.append(data_entry)
    # calculate the average value for each metric
    df = pd.DataFrame(df_data)
    grouped = df.groupby(["Task", "id"])
    sampled_data = []
    task_dict = {}
    for name, group in grouped:
        if name[0] not in task_dict:
            task_dict[name[0]] = 0
        else:
            if task_dict[name[0]] > 5:
                continue
        task_dict[name[0]] += 1
        sampled_data.append(group)

    sampled_df = pd.concat(sampled_data)
    sampled_df.set_index(["Task", "id", "Subtask", "prompt", "Model"], inplace=True)
    # sampled_df = sampled_df.droplevel("tmp")

    sampled_df.to_excel(
        data_folder / "comparison_results.xlsx", index=True, freeze_panes=(1, 4)
    )
    adjust_excel(data_folder / "comparison_results.xlsx")
    logger.info(
        f"Saved comparison results to {data_folder / 'comparison_results.xlsx'}"
    )


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(name)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    parser = argparse.ArgumentParser()
    parser.add_argument("--data_folder", type=Path, required=True)

    args = parser.parse_args()

    generate_failure_cases(args.data_folder, args.data_folder / "results.json")
